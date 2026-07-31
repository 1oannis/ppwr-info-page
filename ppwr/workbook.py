"""Read the PPWR declaration spreadsheet into a plain data structure.

This module knows nothing about language or HTML. It locates the table by
searching for the header cell rather than by row number, so adding articles or
moving the address block in Excel does not break the build.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

HEADER_CELL = "Artikelnummer"
DISTRIBUTOR_MARKER = "In Verkehrbringer:"


class WorkbookError(Exception):
    """The spreadsheet does not have the structure the build expects."""


@dataclass(frozen=True)
class Declaration:
    """The declaration in one language. Cells are text; blanks are ``""``."""

    title: str
    distributor_label: str
    distributor_lines: tuple[str, ...]
    columns: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]


def find_workbook(data_dir: Path) -> Path:
    """Return the single spreadsheet in ``data_dir``.

    Excel lock files (``~$name.xlsx``) are ignored. More than one real
    spreadsheet is an error rather than a guess about which one is current.
    """
    candidates = sorted(
        path for path in data_dir.glob("*.xlsx") if not path.name.startswith("~$")
    )
    if not candidates:
        raise WorkbookError(f"no .xlsx file found in {data_dir}")
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise WorkbookError(
            f"expected exactly one .xlsx in {data_dir}, found {len(candidates)}: {names}"
        )
    return candidates[0]


def _text(value: object, coordinate: str = "") -> str:
    """Return ``value`` as stripped text, or raise if it is not text.

    Excel silently reformats a cell's *display*, not its underlying value:
    an article number typed without its "-01" suffix becomes an int and
    loses its leading zeros, a percent-formatted cell becomes a bare float.
    Reproducing Excel's display formatting is unwinnable, so a non-string
    cell in data read as text is rejected rather than restringified into
    something that quietly does not match what the sheet shows.
    """
    if value is None:
        return ""
    if not isinstance(value, str):
        where = f" at {coordinate}" if coordinate else ""
        raise WorkbookError(
            f"cell{where} contains a {type(value).__name__} ({value!r}), not text - "
            "format the column as Text in Excel so the published value is exactly what you see"
        )
    return value.strip()


def read_declaration(path: Path) -> Declaration:
    """Parse the first worksheet of ``path`` into a :class:`Declaration`."""
    book = openpyxl.load_workbook(path, data_only=True)
    grid = [
        [_text(cell.value, cell.coordinate) for cell in row]
        for row in book.worksheets[0].iter_rows(values_only=False)
    ]

    header_index = next(
        (index for index, row in enumerate(grid) if row and row[0] == HEADER_CELL),
        None,
    )
    if header_index is None:
        raise WorkbookError(
            f"{path.name}: no header row - expected a cell {HEADER_CELL!r} in column A"
        )

    if _distributor_index(grid, header_index) is None:
        raise WorkbookError(
            f"{path.name}: no distributor block - expected a cell {DISTRIBUTOR_MARKER!r} "
            "in column A above the header row"
        )

    header_row = grid[header_index]
    columns: list[str] = []
    gap_index: int | None = None
    for index, cell in enumerate(header_row):
        if not cell:
            gap_index = index
            break
        columns.append(cell)

    if gap_index is not None:
        beyond = [cell for cell in header_row[gap_index + 1:] if cell]
        if beyond:
            gap_column = get_column_letter(gap_index + 1)
            raise WorkbookError(
                f"{path.name}: header row has a blank cell at column {gap_column} "
                f"but {beyond!r} follows it - the header row must not have gaps"
            )

    rows: list[tuple[str, ...]] = []
    blank_row_index: int | None = None
    for offset, row in enumerate(grid[header_index + 1:]):
        if not row or not row[0]:
            blank_row_index = header_index + 1 + offset
            break
        rows.append(tuple(row[i] if i < len(row) else "" for i in range(len(columns))))

    if blank_row_index is not None:
        for index, row in enumerate(grid[blank_row_index + 1:], start=blank_row_index + 1):
            if row and row[0]:
                raise WorkbookError(
                    f"{path.name}: row {index + 1} has an article number but row "
                    f"{blank_row_index + 1} above it is blank - remove the blank row "
                    "or fill in the gap"
                )

    return Declaration(
        title=grid[0][0] if grid and grid[0] else "",
        distributor_label=_distributor_label(grid, header_index),
        distributor_lines=_distributor_lines(grid, header_index),
        columns=tuple(columns),
        rows=tuple(rows),
    )


def _distributor_index(grid: list[list[str]], header_index: int) -> int | None:
    return next(
        (
            index
            for index, row in enumerate(grid[:header_index])
            if row and row[0] == DISTRIBUTOR_MARKER
        ),
        None,
    )


def _distributor_label(grid: list[list[str]], header_index: int) -> str:
    index = _distributor_index(grid, header_index)
    return DISTRIBUTOR_MARKER if index is not None else ""


def _distributor_lines(grid: list[list[str]], header_index: int) -> tuple[str, ...]:
    index = _distributor_index(grid, header_index)
    if index is None:
        return ()
    lines: list[str] = []
    for row in grid[index + 1: header_index]:
        if not row or not row[0]:
            break
        lines.append(row[0])
    return tuple(lines)
